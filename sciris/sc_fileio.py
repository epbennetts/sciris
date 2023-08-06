"""
Functions for reading/writing to files, including pickles, JSONs, and Excel.

Highlights:
    - :func:`sc.save() <save>` / :func:`sc.load() <load>`: efficiently save/load any Python object (via pickling)
    - :func:`sc.savetext() <savetext>` / :func:`sc.loadtext() <loadtext>`: likewise, for text
    - :func:`sc.savejson() <savejson>` / :func:`sc.loadjson() <loadjson>`: likewise, for JSONs
    - :func:`sc.saveyaml() <saveyaml>` / :func:`sc.saveyaml() <saveyaml>`: likewise, for YAML
    - :func:`sc.thisdir() <thisdir>`: get current folder
    - :func:`sc.getfilelist() <getfilelist>`: easy way to access glob
    - :func:`sc.rmpath() <rmpath>`: remove files and folders
"""

##############################################################################
#%% Imports
##############################################################################

# Basic imports
import io
import os
import json
import uuid
import dill
import shutil
import string
import inspect
import importlib
import warnings
import numpy as np
import pandas as pd
import datetime as dt
import gzip as gz
import pickle as pkl
import zstandard as zstd
from zipfile import ZipFile
from contextlib import closing
from pathlib import Path
from glob import glob as pyglob
import fnmatch as fnm
from . import sc_settings as scs
from . import sc_utils as scu
from . import sc_printing as scp
from . import sc_datetime as scd
from . import sc_odict as sco
from . import sc_dataframe as scdf


##############################################################################
#%% Pickling functions
##############################################################################

__all__ = ['load', 'save', 'loadobj', 'saveobj', 'zsave', 'loadstr', 'dumpstr']


# Define common error messages
def _gziperror(filename):
    ''' Base error message when a file couldn't be opened '''
    return f'''
Unable to load
    {filename}
as either a gzipped, zstandard or regular pickle file. Ensure that it is actually a pickle file.
'''

def _unpicklingerror(filename):
    ''' Base error message when a file couldn't be loaded '''
    return f'''
Unable to load file: "{filename}"
as a gzipped pickle file. Loading pickles can fail if Python modules have changed
since the object was saved. If you are loading a custom class that has failed,
you can use the remapping argument; see the sc.load() docstring for details.
Otherwise, you might need to revert to the previous version of that module (e.g.
in a virtual environment), save in a format other than pickle, reload in a different
environment with the new version of that module, and then re-save as a pickle.

For general information on unpickling errors, see e.g.:
https://wiki.python.org/moin/UsingPickle
https://stackoverflow.com/questions/41554738/how-to-load-an-old-pickle-file
'''


def _load_filestr(filename, folder):
    ''' Try different options for loading a file on disk into a string -- not for external use '''
    
    # Handle loading of either filename or file object
    if isinstance(filename, Path):
        filename = str(filename)
    if scu.isstring(filename):
        argtype = 'filename'
        filename = makefilepath(filename=filename, folder=folder, makedirs=False) # If it is a file, validate the folder (but don't create one if it's missing)
    elif isinstance(filename, io.BytesIO):
        argtype = 'fileobj'
    else: # pragma: no cover
        errormsg = f'First argument to sc.load() must be a string or file object, not {type(filename)}; see also sc.loadstr()'
        raise TypeError(errormsg)
    fileargs = {'mode': 'rb', argtype: filename}
    
    try:
        with gz.GzipFile(**fileargs) as fileobj:
            filestr = fileobj.read() # Convert it to a string
    except Exception as E: # pragma: no cover
        exc = type(E) # Figure out what kind of error it is
        if exc == FileNotFoundError: # This is simple, just raise directly
            raise E
        elif exc == gz.BadGzipFile:
            try: # If the gzip file failed, first try as a zstd compressed object
                with open(filename, 'rb') as fh:
                    zdcompressor = zstd.ZstdDecompressor()
                    with zdcompressor.stream_reader(fh) as fileobj:
                        filestr = fileobj.read()
            except Exception as E2: # If that fails...
                try: # Try as a regular binary object
                    with open(filename, 'rb') as fileobj:
                        filestr = fileobj.read() # Convert it to a string
                except Exception as E3:
                    try: # And finally as a regular object
                        with open(filename, 'r') as fileobj:
                            filestr = fileobj.read() # Convert it to a string
                    except Exception as E4:
                        gziperror = _gziperror(filename) + f'\nAdditional errors encountered:\n{str(E2)}\n{str(E3)}\n{str(E4)}'
                        raise exc(gziperror) from E
        else:
            exc = type(E)
            errormsg = 'sc.load(): Could not open the # pragma: no cover file string for an unknown reason; see error above for details'
            raise exc(errormsg) from E
    return filestr


def load(filename=None, folder=None, verbose=False, die=None, remapping=None, 
         method='pickle', auto_remap=True, **kwargs):
    '''
    Load a file that has been saved as a gzipped pickle file, e.g. by :func:`sc.save() <save>`.
    Accepts either a filename (standard usage) or a file object as the first argument.
    Note that :func:`sc.load() <load>`/:func:`sc.loadobj() <loadobj>` are aliases of each other.

    **Note 1**: Since this function relies on pickle, it can potentially execute arbitrary
    code, so you should only use it with sources you trust. For more information, see:
    https://docs.python.org/3/library/pickle.html

    **Note 2**: When a pickle file is loaded, Python imports any modules that are referenced
    in it. This is a problem if module has been renamed. In this case, you can
    use the ``remapping`` argument to point to the new modules or classes. For
    more robustness, use the :func:`sc.savearchive() <savearchive>`/:func:`sc.loadarchive() <loadarchive>`
    functions.


    Args:
        filename   (str/Path): the filename (or full path) to load
        folder     (str/Path): the folder
        verbose    (bool):     print details
        die        (bool):     whether to raise an exception if errors are encountered (otherwise, load as much as possible)
        remapping  (dict):     way of mapping old/unavailable module names to new
        method     (str):      method for loading (usually pickle or dill)
        auto_remap (bool):     whether to use known deprecations to load failed pickles
        kwargs     (dict):     passed to ``pickle.loads()``/``dill.loads()``

    **Examples**::

        obj = sc.load('myfile.obj') # Standard usage
        old = sc.load('my-old-file.obj', method='dill', ignore=True) # Load classes from saved files
        old = sc.load('my-old-file.obj', remapping={'foo.Bar':cat.Mat}) # If loading a saved object containing a reference to foo.Bar that is now cat.Mat
        old = sc.load('my-old-file.obj', remapping={'foo.Bar':('cat', 'Mat')}) # Equivalent to the above

    | *New in version 1.1.0:* "remapping" argument
    | *New in version 1.2.2:* ability to load non-gzipped pickles; support for dill; arguments passed to loader
    '''
    # Load the file
    filestr = _load_filestr(filename, folder)

    # Unpickle it
    try:
        kw = dict(verbose=verbose, die=die, remapping=remapping, method=method, auto_remap=auto_remap)
        obj = _unpickler(filestr, **kw, **kwargs) # Unpickle the data
    except Exception as E: # pragma: no cover
        exc = type(E) # Figure out what kind of error it is
        errormsg = _unpicklingerror(filename) + '\n\nSee the stack trace above for more information on this specific error.'
        raise exc(errormsg) from E

    # If it loaded but with errors, print them here
    if isinstance(obj, Failed):
        print(_unpicklingerror(filename))
    elif verbose: # pragma: no cover
        print(f'Object loaded from "{filename}"')

    return obj


def save(filename='default.obj', obj=None, folder=None, method='pickle', compression='gzip', 
         compresslevel=5, verbose=0, sanitizepath=True, die=False, allow_empty=False, **kwargs):
    '''
    Save any object to disk
    
    This function is similar to :func:`pickle.dump()` in that it serializes the object
    to a file. Key differences include:
        
        - It takes care of opening/closing the file for writing
        - It compresses the data by default
        - It supports different serialization methods (e.g. pickle or dill)
    
    Once an object is saved, it can be loaded with :func:`sc.load() <load>`. Note that 
    :func:`sc.save() <save>`/:func:`sc.saveobj() <saveobj>` are identical.
    
    
    **Note 1**: Since this function relies on pickle, it can potentially execute arbitrary
    code, so you should only use it with sources you trust. For more information, see:
    https://docs.python.org/3/library/pickle.html
        
    **Note 2**: When a pickle file is loaded, Python imports any modules that are referenced
    in it. This is a problem if module has been renamed (in which case the pickle
    usually can't be opened). For more robustness (e.g. to pickle custom classes), use 
    ``method='dill'`` and/or the :func:`sc.savearchive() <savearchive>`/:func:`sc.loadarchive() <loadarchive>`
    functions.
    
    If you do not need to save arbitrary Python and just need to save data, consider
    saving the data in a standard format, e.g. JSON (:func:`sc.savejson() <savejson>`). For large
    amounts of tabular data, also consider formats like HDF5 or PyArrow.

    Args:
        filename      (str/path) : the filename or path to save to; if None, return an io.BytesIO filestream instead of saving to disk
        obj           (anything) : the object to save
        folder        (str)      : optional additional folder, passed to :func:`sc.makepath() <makepath>`
        method        (str)      : whether to use 'pickle' (default) or 'dill'
        compression   (str)      : type of compression to use: 'gzip' (default), 'zstd' (zstandard), or 'none' (no compression)
        compresslevel (int)      : the level of gzip/zstd compression (1 to 9 for gzip, -7 to 22 for zstandard, default 5)
        verbose       (int)      : level of detail to print
        sanitizepath  (bool)     : whether to sanitize the path prior to saving
        die           (bool)     : whether to fail if the object can't be pickled (else, try dill); if die is 'never'
        allow_empty   (bool)     : whether or not to allow "None" to be saved (usually treated as an error)
        kwargs        (dict)     : passed to :func:`pickle.dumps()` (or ``dill.dumps()``)

    **Examples**::

        # Standard usage
        my_obj = ['this', 'is', 'my', 'custom', {'object':44}]
        sc.save('myfile.obj', my_obj)
        loaded = sc.load('myfile.obj')
        assert loaded == my_obj
        
        # Use dill instead, to save custom classes as well
        class MyClass:
            def __init__(self, x):
                self.data = np.random.rand(100) + x
            def sum(self):
                return self.data.sum()
        my_class = MyClass(10)
        
        sc.save('my_class.obj', my_class, method='dill', compression='zstd')
        loaded = sc.load('my_class.obj') # With dill, can be loaded anywhere, not just in the same script
        assert loaded.sum() == my_class.sum()
    
    See also :func:`sc.zsave() <zsave>` (identical except defaults to zstandard compression).

    | *New in version 1.1.1:* removed Python 2 support.
    | *New in version 1.2.2:* automatic swapping of arguments if order is incorrect; correct passing of arguments
    | *New in version 2.0.4:* "die" argument for saving as dill
    | *New in version 2.1.0:* "zstandard" compression method
    | *New in version 3.0.0:* "allow_empty" argument; removed "args"
    '''

    def serialize(fileobj, obj, success, **kwargs):
        ''' Actually write a serial bytestream to disk '''
        # Try pickle first
        if method == 'pickle':
            try:
                if verbose>=2: print('Saving as pickle...')
                _savepickle(fileobj, obj, **kwargs) # Use pickle
                success = True
            except Exception as E: # pragma: no cover
                if die is True:
                    raise E
                else:
                    if verbose>=2: print(f'Exception when saving as pickle ({repr(E)}), saving as dill...')

        # If dill is requested or pickle failed, use dill
        if not success: # pragma: no cover
            if verbose>=2: print('Saving as dill...')
            _savedill(fileobj, obj, **kwargs)
        
        return
    
    # Handle path
    if filename is None: # If the user explicitly specifies None as the file, create a byte stream instead
        tobytes = True
        bytestream = io.BytesIO()
    else:
        tobytes = False
        bytestream = None

        # Process and sanitize the filename passed in by the user
        filetypes = (str, type(Path()), type(None))
        if isinstance(filename, Path): # If it's a path object, convert to string
            filename = str(filename)

        if not isinstance(filename, filetypes): # pragma: no cover
            if isinstance(obj, filetypes):
                print(f'Warning: filename was not supplied as a valid type ({type(filename)}) but the object was ({type(obj)}); automatically swapping order')
                real_obj = filename
                real_file = obj
                filename = real_file
                obj = real_obj
            else:
                errormsg = f'Filename type {type(filename)} is not valid: must be one of {filetypes}'
                raise Exception(errormsg)

        filename = makefilepath(filename=filename, folder=folder, sanitize=sanitizepath, makedirs=True)

    # Handle object
    if obj is None: # pragma: no cover
        errormsg = "No object was supplied to saveobj(), or the object was empty; if this is intentional, set die='never'"
        if not allow_empty and die != 'never': # die = 'never' is kept for backwards compatibility
            raise ValueError(errormsg)

    # Compress and actually save
    success = False
    
    if compression in ['gz', 'gzip']:  # Main use case
        # File extension is .gz
        with gz.GzipFile(filename=filename, fileobj=bytestream, mode='wb', compresslevel=compresslevel) as fileobj:
            serialize(fileobj, obj, success, **kwargs)
    
    else:
        if tobytes: # pragma: no cover
            filecontext = closing(bytestream)
        else:
            filecontext = open(filename, 'wb')
        
        if compression in ['zst', 'zstd', 'zstandard']:
            # File extension is .zst
            with filecontext as fh:
                zcompressor = zstd.ZstdCompressor(level=compresslevel)
                with zcompressor.stream_writer(fh) as fileobj:
                    serialize(fileobj, obj, success, **kwargs)
        elif compression in ['none']:
            # File extension can be anything
            with filecontext as fileobj:
                serialize(fileobj, obj, success, **kwargs)
        else: # pragma: no cover
            errormsg = f"Invalid compression format '{compression}': must be 'gzip', 'zstd', or 'none'"
            raise ValueError(errormsg)

    if verbose and filename: # pragma: no cover
        print(f'Object saved to "{filename}"')

    if filename:
        return filename
    else: # pragma: no cover
        bytestream.seek(0)
        return bytestream


# Backwards compatibility for core functions
loadobj = load
saveobj = save


def zsave(*args, compression='zstd', **kwargs):
    '''
    Save a file using zstandard (instead of gzip) compression. This is an alias
    for ``sc.save(..., compression='zstd')``; see :func:`sc.save() <save>` for details.
    
    Note: there is no matching function ``sc.zload()`` since :func:`sc.load() <load>` will
    automatically try loading zstandard.
    
    *New in version 2.1.0.*
    '''
    return save(*args, compression=compression, **kwargs)


def loadstr(string, **kwargs):
    '''
    Like :func:`sc.load() <load>`, but for a bytes-like string (rarely used).
    
    Args:
        string (str): the bytes-like string to load
        kwargs (dict): passed to :func:`sc.load() <load>`

    **Example**::

        obj = sc.objdict(a=1, b=2)
        bytestring = sc.dumpstr(obj)
        obj2 = sc.loadstr(bytestring)
        assert obj == obj2
    
    | *New in version 3.0.0:* uses :func:`sc.load() <load>` for more robustness
    '''
    with closing(io.BytesIO(string)) as bytestream: # Open a "fake file" with the Gzip string pickle in it
        obj = load(bytestream, **kwargs)
    return obj


def dumpstr(obj=None, **kwargs):
    '''
    Dump an object to a bytes-like string (rarely used by the user); see :func:`sc.save() <save>`
    instead.
    
    Args:
        obj (any): the object to convert
        kwargs (dict): passed to :func:`sc.save() <save>`
    
    *New in version 3.0.0:* uses :func:`sc.save() <save>` for more robustness
    '''
    bytesobj = save(filename=None, obj=obj, **kwargs)
    result = bytesobj.read() # Read all of the content into result
    return result


##############################################################################
#%% Other file functions
##############################################################################

__all__ += ['loadtext', 'savetext', 'loadzip', 'unzip', 'savezip', 'path', 'ispath', 
            'thisfile', 'thisdir', 'thispath', 'getfilelist', 'glob', 'getfilepaths', 
            'sanitizefilename', 'sanitizepath', 'makefilepath', 'makepath', 'rmpath']


def loadtext(filename=None, folder=None, splitlines=False):
    '''
    Convenience function for reading a text file

    **Example**::

        mytext = sc.loadtext('my-document.txt')
    '''
    filename = makefilepath(filename=filename, folder=folder)
    with open(filename) as f:
        output = f.read()
    if splitlines:
        output = output.splitlines()
    return output


def savetext(filename=None, string=None, **kwargs):
    '''
    Convenience function for saving a text file -- accepts a string or list of strings;
    can also save an arbitrary object, in which case it will first convert to a string.
    
    Args:
        filename (str): the filename to save to
        string (str): the string (or object) to save
        kwargs (dict): passed to :func:`np.savetxt() <numpy.savetxt>` if saving an array

    **Example**::

        text = ['Here', 'is', 'a', 'poem']
        sc.savetext('my-poem.txt', text)
    '''
    is_array = scu.isarray(string)
    if isinstance(string, list):
        string = '\n'.join(str(string)) # Convert from list to string)
    elif not is_array and not scu.isstring(string):
        string = str(string)
    filename = makefilepath(filename=filename, makedirs=True)
    if is_array: # Shortcut to Numpy for saving arrays -- basic CSV
        kw = scu.mergedicts(dict(fmt='%s', delimiter=', '), kwargs)
        np.savetxt(filename, string, **kw)
    else: # Main use case: save text
        with open(filename, 'w') as f:
            f.write(string)
    return


def loadzip(filename=None, folder=None, **kwargs):
    '''
    Load the contents of a zip file into a variable

    Args:
        filename (str/path): the name of the zip file to write to
        folder (str): optional additional folder for the filename
        kwargs (dict): passed to :func:`sc.load() <load>`
    
    Returns:
        dict with each file loaded as a key

    **Example**::

        data = sc.loadzip('my-files.zip')

    | *New in version 2.0.0.*
    | *New in version 3.0.0:* load into memory instead of extracting to disk; see :func:`sc.unzip() <unzip>` for extracting
    '''
    filename = makefilepath(filename=filename, folder=folder)
    output = dict()
    with ZipFile(filename, 'r') as zf: # Create the zip file
        names = zf.namelist()
        for name in names:
            val = zf.read(name)
            try:
                val = loadstr(val, **kwargs) # Try to load as a pickle or some kind of valid file
            except:
                pass # Otherwise, just return the raw value
            output[name] = val
    return output


def unzip(filename=None, outfolder='.', folder=None, members=None):
    '''
    Convenience function for reading a zip file

    Args:
        filename (str/path): the name of the zip file to write to
        outfolder (str/path): the path location to extract the files to (default: current folder)
        folder (str): optional additional folder for the filename
        members (list): optional list of members
    
    Returns:
        list of the names of the unzipped files

    **Example**::

        sc.unzip('my-files.zip', outfolder='my_data') # extracts all files 

    | *New in version 3.0.0* (equivalent to sc.loadzip(..., extract=True) previously)
    '''
    filename = makefilepath(filename=filename, folder=folder)
    with ZipFile(filename, 'r') as zf: # Load the zip file
        names = zf.namelist()
        zf.extractall(outfolder, members=members)
    
    output = [makefilepath(filename=name, folder=outfolder, makedirs=True) for name in names]
    return output


def savezip(filename=None, files=None, data=None, folder=None, sanitizepath=True, 
            basename=False, tobytes=True, verbose=True, **kwargs):
    '''
    Create a zip file from the supplied list of files (or less commonly, supplied data)

    Args:
        filename (str/path): the name of the zip file to write to
        files (list): file(s) and/or folder(s) to compress
        data (dict): if supplied, write this data as well or instead (must be a dictionary of filename keys and data values)
        folder (str): optional additional folder for the filename
        sanitizepath (bool): whether to sanitize the path prior to saving
        basename (bool): whether to use only the file's basename as the name inside the zip file (otherwise, store folder info)
        tobytes (bool): if data is provided, convert it automatically to bytes (otherwise, up to the user)
        verbose (bool): whether to print progress
        kwargs (dict): passed to :func:`sc.save() <save>`

    **Examples**::

        scripts = sc.getfilelist('./code/*.py')
        sc.savezip('scripts.zip', scripts)

        sc.savezip('mydata.zip', data=dict(var1='test', var2=np.random.rand(3)))

    | *New in version 2.0.0:* saving data
    | *New in version 3.0.0:* "tobytes" argument and kwargs; "filelist" renamed "files"
    '''

    # Handle inputs
    fullpath = makefilepath(filename=filename, folder=folder, sanitize=sanitizepath, makedirs=True)
    
    
    # If data is provided, do simple validation
    if data is not None:
        if not isinstance(data, dict): # pragma: no cover
            errormsg = 'Data has invalid format: must be a dictionary of filename keys and data values'
            raise ValueError(errormsg)
    else:
        data = {}
    
    # Typical use case: data is not provided, handle files
    if files:
        origfilelist = [path(file) for file in scu.tolist(files)]
        
        # Handle subfolders
        extfilelist = scu.dcp(origfilelist) # An extended file list, including recursion into subfolders
        for orig in origfilelist:
            if orig.is_dir():
                contents = getfilelist(orig, abspath=False, recursive=True, aspath=True)
                extfilelist.extend(contents[1:]) # Skip the first entry since it's the folder that's already in the list
        
        # Remove duplicates to create the final file list
        filelist = []
        for efile in extfilelist:
            if efile not in filelist:
                filelist.append(efile)
            
    # Write zip file
    with ZipFile(fullpath, 'w') as zf: # Create the zip file
        if data: # Optionally also save data
            for key,val in data.items():
                if tobytes:
                    val = dumpstr(val, **kwargs)
                zf.writestr(key, val)
        if files: # Main use case, save files
            for thisfile in filelist:
                thispath = makefilepath(filename=thisfile, abspath=False, makedirs=False)
                thisname = os.path.basename(thisfile) if basename else thisfile
                zf.write(thispath, thisname) # Actually save
        if not data and not files:
            errormsg = 'No data and no files provided: nothing to save!'
            raise FileNotFoundError(errormsg)
        
    if verbose: print(f'Zip file saved to "{fullpath}"')
    return fullpath


def path(*args, **kwargs):
    '''
    Alias to ``pathlib.Path()`` with some additional input sanitization:

        - ``None`` entries are removed
        - a list of arguments is converted to separate arguments
        
    **Examples**::
        
        sc.path('thisfile.py') # Returns PosixPath('thisfile.py')
        
        sc.path('/a/folder', None, 'a_file.txt') # Returns PosixPath('/a/folder/a_file.txt')

    | *New in version 1.2.2.*
    | *New in version 2.0.0:* handle None or list arguments
    '''

    # Handle inputs
    new_args = []
    for arg in args:
        if isinstance(arg, list):
            new_args.extend(arg)
        else:
            new_args.append(arg)
    new_args = [arg for arg in new_args if arg is not None]

    # Create the path
    output = Path(*new_args, **kwargs)

    return output

path.__doc__ += '\n\n' + Path.__doc__


def ispath(obj):
    '''
    Alias to isinstance(obj, Path).

    *New in version 2.0.0.*
    '''
    return isinstance(obj, Path)


def thisfile(frame=1, aspath=None):
    '''
    Return the full path of the current file.
    
    Args:
        frame (int): which frame to pull the filename from (default 1, the file that calls this function)
        aspath (bool): whether to return a Path object
    
    **Examples**::
        
        my_script_name = sc.thisfile() # Get the name of the current file
        calling_script = sc.thisfile(frame=2) # Get the name of the script that called this script

    *New in verison 2.1.0.*
    '''
    if aspath is None: aspath = scs.options.aspath
    file = inspect.stack()[frame][1] # Adopted from Atomica
    if aspath:
        file = Path(file)
    return file


def thisdir(file=None, path=None, *args, frame=1, aspath=None, **kwargs):
    '''
    Tiny helper function to get the folder for a file, usually the current file.
    If not supplied, then use the current file.

    Args:
        file (str): the file to get the directory from; usually __file__
        path (str/list): additional path to append; passed to os.path.join()
        args  (list): also passed to os.path.join()
        frame (int): if file is None, which frame to pull the folder from (default 1, the file that calls this function)
        aspath (bool): whether to return a Path object instead of a string
        kwargs (dict): passed to Path()

    Returns:
        filepath (str): the full path to the folder (or filename if additional arguments are given)

    **Examples**::

        thisdir = sc.thisdir() # Get folder of calling file
        thisdir = sc.thisdir('.') # Ditto (usually)
        thisdir = sc.thisdir(__file__) # Ditto (usually)
        file_in_same_dir = sc.thisdir(path='new_file.txt')
        file_in_sub_dir = sc.thisdir('..', 'tests', 'mytests.py') # Merge parent folder with sufolders and a file
        np_dir = sc.thisdir(np) # Get the folder that Numpy is loaded from (assuming "import numpy as np")

    | *New in version 1.1.0:* "as_path" argument renamed "aspath"
    | *New in version 1.2.2:* "path" argument
    | *New in version 1.3.0:* allow modules
    | *New in version 2.1.0:* frame argument
    '''
    if file is None: # No file: use the current folder
        if scu.isjupyter():
            file = os.path.abspath(os.path.expanduser('file_placeholder')) # This is as best we can do on Jupyter
        else:
            file = thisfile(frame=frame+1) # Need +1 since want the calling script
    elif hasattr(file, '__file__'): # It's actually a module
        file = file.__file__
    if aspath is None: aspath = scs.options.aspath
    folder = os.path.abspath(os.path.dirname(file))
    path = scu.mergelists(path, *args)
    filepath = os.path.join(folder, *path)
    if aspath:
        filepath = Path(filepath, **kwargs)
    return filepath


def thispath(*args, frame=1, aspath=True, **kwargs):
    '''
    Alias for :func:`sc.thisdir() <thisdir>` that returns a path by default instead of a string.
    
    *New in version 2.1.0.*
    '''
    return thisdir(*args, frame=frame+1, aspath=aspath, **kwargs)

thispath.__doc__ += '\n\n' + thisdir.__doc__


def getfilelist(folder=None, pattern=None, fnmatch=None, abspath=False, nopath=False, 
                filesonly=False, foldersonly=False, recursive=True, aspath=None):
    '''
    A shortcut for using :func:`glob.glob()`.
    
    Note that :func:`sc.getfilelist() <getfilelist>` and :func:`sc.glob() <glob>` 
    are aliases of each other.

    Args:
        folder      (str):  the folder to find files in (default, current)
        pattern     (str):  the pattern to match (default, wildcard); can be excluded if part of the folder
        fnmatch     (str):  optional additional string to filter results by
        abspath     (bool): whether to return the full path
        nopath      (bool): whether to return no path
        filesonly   (bool): whether to only return files (not folders)
        foldersonly (bool): whether to only return folders (not files)
        recursive   (bool): passed to :func:`glob.glob()` (note: ** is required as the pattern to match all subfolders)
        aspath      (bool): whether to return Path objects
        
    Returns:
        List of files/folders

    **Examples**::

        sc.getfilelist() # return all files and folders in current folder
        sc.getfilelist('~/temp', '*.py', abspath=True) # return absolute paths of all Python files in ~/temp folder
        sc.getfilelist('~/temp/*.py') # Like above
        sc.getfilelist(fnmatch='*.py') # Recursively find all files ending in .py

    | *New in version 1.1.0:* "aspath" argument
    | *New in version 2.1.0:* default pattern of "**"; "fnmatch" argument; default recursive=True
    '''
    if folder is None:
        folder = '.'
    folder = os.path.expanduser(folder)
    if abspath:
        folder = os.path.abspath(folder)
    if os.path.isdir(folder) and pattern is None:
        pattern = '**'
    if aspath is None: 
        aspath = scs.options.aspath
    globstr = os.path.join(folder, pattern) if pattern else folder
    filelist = sorted(pyglob(globstr, recursive=recursive))
    if filesonly:
        filelist = [f for f in filelist if os.path.isfile(f)]
    elif foldersonly:
        filelist = [f for f in filelist if os.path.isdir(f)]
    if nopath:
        filelist = [os.path.basename(f) for f in filelist]
    if fnmatch:
        filelist = [f for f in filelist if fnm.fnmatch(f, fnmatch)]
    if aspath:
        filelist = [Path(f) for f in filelist]
    return filelist


# Define as an alias
glob = getfilelist


def getfilepaths(*args, aspath=True, **kwargs):
    '''
    Alias for :func:`sc.getfilelist() <getfilelist>` that returns paths by default instead of strings.
    
    *New version 2.1.0.*
    '''
    return getfilelist(*args, aspath=True, **kwargs)

getfilepaths.__doc__ += '\n\n' + getfilelist.__doc__


def sanitizefilename(filename, sub='_', allowspaces=False, asciify=True, strict=False, disallowed=None, aspath=False):
    '''
    Takes a potentially Linux- and Windows-unfriendly candidate file name, and
    returns a "sanitized" version that is more usable.
    
    Args:
        filename (str): the filename to sanitize
        sub (str): the character to substitute unsafe input characters with
        allowspaces (bool): whether to allow spaces in the filename
        asciify (bool): whether to convert the string from Unicode to ASCII
        strict (bool): whether to remove (almost) all non-alphanumeric characters
        disallowed (str): optionally supply a custom list of disallowed characters
        aspath (bool): whether to return a Path object

    **Example**::

        bad = 'Nöt*a   file&name?!.doc'
        good = sc.sanitizefilename(bad)
        
    *New version 2.0.1:* arguments "sub", "allowspaces", "asciify", "strict", and "disallowed"
    '''
    
    # Handle options
    if asciify:
        filename = scu.asciify(filename) # Ensure it's ASCII compatible
    if disallowed is None:
        if strict:
            disallowed = '''!"#$%&\'()*+,/:;<=>?@[\\]^`{|}~\t\n\r\x0b\x0c'''
        else:
            disallowed = '''\\/:*?!"'<>|'''
    if not allowspaces:
        disallowed += ' '
    
    # Create the filename
    sanitized = ''
    for letter in filename:
        if letter in string.printable or not asciify:
            if letter in disallowed:
                sanitized += sub
            else:
                sanitized += letter
    
    if aspath is None: 
        aspath = scs.options.aspath
    if aspath:
        sanitized = Path(sanitized)

    return sanitized # Return the sanitized file name.


def sanitizepath(*args, aspath=True, **kwargs):
    '''
    Alias for :func:`sc.sanitizefilename() <sanitizefilename>` that returns a path by default instead of a string.
    
    *New version 2.1.0.*
    '''
    return sanitizefilename(*args, aspath=True, **kwargs)

sanitizepath.__doc__ += '\n\n' + sanitizefilename.__doc__



def makefilepath(filename=None, folder=None, ext=None, default=None, split=False, aspath=None, abspath=True, makedirs=False, checkexists=None, sanitize=False, die=True, verbose=False):
    '''
    Utility for taking a filename and folder -- or not -- and generating a
    valid path from them. By default, this function will combine a filename and
    folder using os.path.join, create the folder(s) if needed with os.makedirs,
    and return the absolute path.

    Args:
        filename    (str or Path)   : the filename, or full file path, to save to -- in which case this utility does nothing
        folder      (str/Path/list) : the name of the folder to be prepended to the filename; if a list, fed to :func:`os.path.join()`
        ext         (str)           : the extension to ensure the file has
        default     (str or list)   : a name or list of names to use if filename is None
        split       (bool)          : whether to return the path and filename separately
        aspath      (bool)          : whether to return a Path object (default: set by ``sc.options.aspath``)
        abspath     (bool)          : whether to conver to absolute path
        makedirs    (bool)          : whether or not to make the folders to save into if they don't exist
        checkexists (bool)          : if False/True, raises an exception if the path does/doesn't exist
        sanitize    (bool)          : whether or not to remove special characters from the path; see :func:`sc.sanitizepath() <sanitizepath>` for details
        die         (bool)          : whether or not to raise an exception if cannot create directory failed (otherwise, return a string)
        verbose     (bool)          : how much detail to print

    Returns:
        filepath (str or Path): the validated path (or the folder and filename if split=True)

    **Simple example**::

        filepath = sc.makefilepath('myfile.obj') # Equivalent to os.path.abspath(os.path.expanduser('myfile.obj'))

    **Complex example**::

        filepath = makefilepath(filename=None, folder='./congee', ext='prj', default=[project.filename, project.name], split=True, abspath=True, makedirs=True)

    Assuming project.filename is None and project.name is "recipe" and ./congee
    doesn't exist, this will makes folder ./congee and returns e.g. ('/home/myname/congee', 'recipe.prj')

    | *New in version 1.1.0:* "aspath" argument
    | *New in version 3.0.0:* "makedirs" defaults to False
    '''

    # Initialize
    filefolder = '' # The folder the file will be located in
    filebasename = '' # The filename
    if aspath is None: aspath = scs.options.aspath

    if isinstance(filename, Path):
        filename = str(filename)
    if isinstance(folder, Path): # If it's a path object, convert to string
        folder = str(folder)
    if isinstance(folder, list): # It's a list, join together # pragma: no cover
        folder = os.path.join(*folder)

    # Process filename
    if filename is None: # pragma: no cover
        defaultnames = scu.tolist(default) # Loop over list of default names
        for defaultname in defaultnames:
            if not filename and defaultname: filename = defaultname # Replace empty name with default name
    if filename is not None: # If filename exists by now, use it
        filebasename = os.path.basename(filename)
        filefolder = os.path.dirname(filename)
    if not filebasename: filebasename = 'default' # If all else fails

    # Add extension if it's defined but missing from the filebasename
    if ext and not filebasename.endswith(ext): # pragma: no cover
        filebasename += '.'+ext
    if verbose: # pragma: no cover
        print(f'From filename="{filename}", default="{default}", extension="{ext}", made basename "{filebasename}"')

    # Sanitize base filename
    if sanitize: filebasename = sanitizefilename(filebasename)

    # Process folder
    if folder is not None: # Replace with specified folder, if defined
        filefolder = folder
    if abspath: # Convert to absolute path
        filefolder = os.path.abspath(os.path.expanduser(filefolder))

    # Make sure folder exists
    if makedirs:
        try:
            os.makedirs(filefolder, exist_ok=True)
        except Exception as E: # pragma: no cover
            if die:
                raise E
            else:
                print(f'Could not create folders: {str(E)}')

    # Create the full path
    filepath = os.path.join(filefolder, filebasename)

    # Optionally check if it exists
    if checkexists is not None: # pragma: no cover
        exists = os.path.exists(filepath)
        errormsg = ''
        if exists and not checkexists:
            errormsg = f'File {filepath} should not exist, but it does'
            if die:
                raise FileExistsError(errormsg)
        if not exists and checkexists:
            errormsg = f'File {filepath} should exist, but it does not'
            if die:
                raise FileNotFoundError(errormsg)
        if errormsg:
            print(errormsg)

    # Decide on output
    if verbose: # pragma: no cover
        print(f'From filename="{filename}", folder="{folder}", made path name "{filepath}"')
    if split: # pragma: no cover
        output = filefolder, filebasename
    elif aspath:
        output = Path(filepath)
    else:
        output = filepath

    return output


def makepath(*args, aspath=True, **kwargs):
    '''
    Alias for :func:`sc.makefilepath() <makefilepath>` that returns a path by default instead of a string
    (with apologies for the confusing terminology, kept for backwards compatibility).
    
    *New version 2.1.0.*
    '''
    return makefilepath(*args, **kwargs, aspath=True)

makepath.__doc__ += '\n\n' + makefilepath.__doc__


def rmpath(path=None, *args, die=True, verbose=True, interactive=False, **kwargs):
    """
    Remove file(s) and folder(s). Alias to :func:`os.remove()` (for files) and :func:`shutil.rmtree()`
    (for folders).

    Arguments:
        path (str/Path/list): file, folder, or list to remove
        args (list): additional paths to remove
        die (bool): whether or not to raise an exception if cannot remove
        verbose (bool): how much detail to print
        interactive (bool): whether to confirm prior to each deletion
        kwargs (dict): passed to :func:`os.remove()`/:func:`shutil.rmtree()`

    **Examples**::

        sc.rmpath('myobj.obj') # Remove a single file
        sc.rmpath('myobj1.obj', 'myobj2.obj', 'myobj3.obj') # Remove multiple files
        sc.rmpath(['myobj.obj', 'tests']) # Remove a file and a folder interactively
        sc.rmpath(sc.getfilelist('tests/*.obj')) # Example of removing multiple files

    *New version 2.0.0.*
    """

    paths = scu.mergelists(path, *args)
    for path in paths:
        if not os.path.exists(path): # pragma: no cover
            errormsg = f'Path "{path}" does not exist'
            if die:
                raise FileNotFoundError(errormsg)
            elif verbose:
                print(errormsg)
        else:
            if os.path.isfile(path):
                rm_func = os.remove
            elif os.path.isdir(path): # pragma: no cover
                rm_func = shutil.rmtree
            else: # pragma: no cover
                errormsg = f'Path "{path}" exists, but is neither a file nor a folder: unable to remove'
                if die:
                    raise FileNotFoundError(errormsg)
                elif verbose:
                    print(errormsg)

        if interactive: # pragma: no cover
            ans = input(f'Remove "{path}"? (y/[n]) ')
            if ans != 'y':
                print(f'  Skipping "{path}"')
                continue

        try:
            rm_func(path)
            if verbose or interactive:
                print(f'Removed "{path}"')
        except Exception as E: # pragma: no cover
            if die:
                raise E
            elif verbose:
                errormsg = f'Could not remove "{path}": {str(E)}'
                print(errormsg)

    return


##############################################################################
#%% JSON functions
##############################################################################

__all__ += ['sanitizejson', 'jsonify', 'printjson', 'readjson', 'loadjson', 'savejson', 
            'readyaml', 'loadyaml', 'saveyaml', 'jsonpickle', 'jsonunpickle']


def jsonify(obj, verbose=True, die=False, tostring=False, **kwargs):
    """
    This is the main conversion function for Python data-structures into JSON-compatible
    data structures (note: :func:`sc.sanitizejson() <sanitizejson>`/:func:`sc.jsonify() <jsonify>` are identical).

    Args:
        obj      (any):  almost any kind of data structure that is a combination of list, :obj:`numpy.ndarray`, odicts, etc.
        verbose  (bool): level of detail to print
        die      (bool): whether or not to raise an exception if conversion failed (otherwise, return a string)
        tostring (bool): whether to return a string representation of the sanitized object instead of the object itself
        kwargs   (dict): passed to json.dumps() if tostring=True

    Returns:
        object (any or str): the converted object that should be JSON compatible, or its representation as a string if tostring=True

    **Example**::
        
        data = dict(a=np.random.rand(3), b=dict(foo='cat', bar='dog'))
        json = sc.jsonify(data)
        jsonstr = sc.jsonify(data, tostring=True, indent=2)
    """
    if obj is None: # Return None unchanged
        output = None

    elif isinstance(obj, (bool, np.bool_)): # It's true/false
        output = bool(obj)

    elif scu.isnumber(obj): # It's a number
        if np.isnan(obj): # It's nan, so return None # pragma: no cover
            output = None
        else:
            if isinstance(obj, (int, np.int64)):
                output = int(obj) # It's an integer
            else:
                output = float(obj)# It's something else, treat it as a float

    elif scu.isstring(obj): # It's a string of some kind
        try:    string = str(obj) # Try to convert it to ascii
        except: string = obj # Give up and use original
        output = string

    elif isinstance(obj, np.ndarray): # It's an array, iterate recursively
        if obj.shape: output = [sanitizejson(p) for p in list(obj)] # Handle most cases, incluing e.g. array([5])
        else:         output = [sanitizejson(p) for p in list(np.array([obj]))] # Handle the special case of e.g. array(5) # pragma: no cover

    elif isinstance(obj, (list, set, tuple)): # It's another kind of interable, so iterate recurisevly
        output = [sanitizejson(p) for p in list(obj)]

    elif isinstance(obj, dict): # It's a dictionary, so iterate over the items
        output = {str(key):sanitizejson(val) for key,val in obj.items()}

    elif isinstance(obj, (dt.time, dt.date, dt.datetime)): # pragma: no cover
        output = str(obj)

    elif isinstance(obj, uuid.UUID): # pragma: no cover
        output = str(obj)

    elif callable(getattr(obj, 'to_dict', None)): # Handle e.g. pandas, where we want to return the object, not the string # pragma: no cover
        output = obj.to_dict()

    elif callable(getattr(obj, 'to_json', None)): # pragma: no cover
        output = obj.to_json()

    elif callable(getattr(obj, 'toJSON', None)): # pragma: no cover
        output = obj.toJSON()

    else: # None of the above
        try:
            output = jsonify(obj.__dict__) # Try converting the contents to JSON
            output = scu.mergedicts({'python_class': str(type(obj))}, output)
        except:
            try:
                output = jsonpickle(obj)
            except Exception as E: # pragma: no cover
                errormsg = f'Could not sanitize "{obj}" {type(obj)} ({E}), converting to string instead'
                if die:       raise TypeError(errormsg)
                elif verbose: print(errormsg)
                output = str(obj)

    # Convert to string if desired
    if tostring:
        output = json.dumps(output, **kwargs)

    return output

# Define alias
sanitizejson = jsonify


def printjson(obj, indent=2, **kwargs):
    '''
    Print an object as a JSON
    
    Acts as an alias to :func:`print(sc.jsonify(..., tostring=True)) <jsonify>`.
    
    Args:
        obj (any): the object to print
        indent (int): the level of indent to use 
        kwargs (dict): passed to :func:`sc.jsonify() <jsonify>`
    
    **Example**::
        
        data = dict(a=dict(x=[1,2,3], y=[4,5,6]), b=dict(foo='string', bar='other_string'))
        sc.printjson(data)
    
    *New in version 3.0.0.*
    '''
    json = jsonify(obj, tostring=True, indent=indent, **kwargs)
    print(json)
    return


def readjson(string, **kwargs):
    '''
    Read JSON from a string
    
    Alias to :func:`json.loads()`.

    Args:
        string (str): a string representation of the JSON
        kwargs (dict): passed to :func:`json.loads()`
    
    See also :func:`sc.loadjson() <loadjson>` for loading a JSON from
    a file.
    
    **Example**::
        
        string = '{"this":1, "is":2, "a":3, "JSON":4}'
        json = sc.readjson(string)
        
    *New in version 3.0.0.*
    '''
    return json.loads(string, **kwargs)


def loadjson(filename=None, folder=None, string=None, fromfile=True, **kwargs):
    '''
    Convenience function for reading a JSON file (or string).

    Args:
        filename (str): the file to load, or the JSON object if using positional arguments
        folder (str): folder if not part of the filename
        string (str): if not loading from a file, a string representation of the JSON
        fromfile (bool): whether or not to load from file
        kwargs (dict): passed to :func:`json.load()`

    Returns:
        output (dict): the JSON object

    **Examples**::

        json = sc.loadjson('my-file.json')
        json = sc.loadjson(string='{"a":null, "b":[1,2,3]}')
    
    See also :func:`sc.readjson() <readjson>` for loading a JSON from
    a string.
    '''
    if string is not None or not fromfile:
        if string is None and filename is not None: # pragma: no cover
            string = filename # Swap arguments
        output = json.loads(string, **kwargs)
    else:
        filepath = makefilepath(filename=filename, folder=folder)
        try:
            with open(filepath) as f:
                output = json.load(f, **kwargs)
        except FileNotFoundError as E: # pragma: no cover
            errormsg = f'No such file "{filename}". Use "string" argument or "fromfile=False" if loading a JSON string rather than a file.'
            raise FileNotFoundError(errormsg) from E
    return output


def savejson(filename=None, obj=None, folder=None, die=True, indent=2, keepnone=False, sanitizepath=True, **kwargs):
    '''
    Convenience function for saving to a JSON file.

    Args:
        filename (str): the file to save
        obj (anything): the object to save; if not already in JSON format, conversion will be attempted
        folder (str): folder if not part of the filename
        die (bool): whether or not to raise an exception if saving an empty object
        indent (int): indentation to use for saved JSON
        keepnone (bool): allow :func:`sc.savejson(None) <savejson(None>` to return 'null' rather than raising an exception
        sanitizepath (bool): whether to sanitize the path prior to saving
        kwargs (dict): passed to :func:`json.dump()`

    Returns:
        The filename saved to

    **Example**::

        json = {'foo':'bar', 'data':[1,2,3]}
        sc.savejson('my-file.json', json)
    
    See also :func:`sc.jsonify() <jsonify>`.
    '''

    filename = makefilepath(filename=filename, folder=folder, sanitize=sanitizepath, makedirs=True)

    if obj is None and not keepnone: # pragma: no cover
        errormsg = 'No object was supplied to savejson(), or the object was empty'
        if die: raise ValueError(errormsg)
        else:   print(errormsg)

    with open(filename, 'w') as f:
        json.dump(sanitizejson(obj), f, indent=indent, **kwargs)

    return filename


def readyaml(string, **kwargs):
    '''
    Read YAML from a string
    
    Alias to :func:`sc.loadyaml(string=...) <loadyaml>`.

    Args:
        string (str): a string representation of the YAML
        kwargs (dict): passed to :func:`sc.loadyaml() <loadyaml>`
    
    See also :func:`sc.loadyaml() <loadyaml>` for loading a YAML from
    a file.
    
    **Example**::
    
        string = '{"this":1, "is":2, "a":3, "YAML":4} # YAML allows comments!'
        yaml = sc.readyaml(string)
    
    *New in version 3.0.0.*
    '''
    return loadyaml(string=string, **kwargs)


def loadyaml(filename=None, folder=None, string=None, fromfile=True, safe=False, loader=None):
    '''
    Convenience function for reading a YAML file (or string).

    Args:
        filename (str): the file to load, or the YAML object if using positional arguments
        folder (str): folder if not part of the filename
        string (str): if not loading from a file, a string representation of the YAML
        fromfile (bool): whether or not to load from file
        safe (bool): whether to use the safe loader
        loader (Loader): custom YAML loader (takes precedence over ``safe``)

    Returns:
        output (dict): the YAML object

    **Examples**::

        yaml = sc.loadyaml('my-file.yaml')
        yaml = sc.loadyaml(string='{"a":null, "b":[1,2,3]}')
    '''
    import yaml # Optional import

    if loader is None:
        if safe: loader = yaml.loader.SafeLoader
        else:    loader = yaml.loader.UnsafeLoader

    if string is not None or not fromfile: # pragma: no cover
        if string is None and filename is not None:
            string = filename # Swap arguments
        output = yaml.load_all(string, loader)
        output = list(output)
    else:
        filepath = makefilepath(filename=filename, folder=folder)
        try:
            with open(filepath) as f:
                output = yaml.load_all(f, loader)
                output = list(output)
        except FileNotFoundError as E: # pragma: no cover
            errormsg = f'No such file "{filename}". Use fromfile=False if loading a YAML string rather than a file.'
            raise FileNotFoundError(errormsg) from E

    # If only a single page, return it directly
    if len(output) == 1:
        output = output[0]

    return output


def saveyaml(filename=None, obj=None, folder=None, die=True, keepnone=False, dumpall=False, sanitizepath=True, **kwargs):
    '''
    Convenience function for saving to a YAML file.

    Args:
        filename (str): the file to save (if empty, return string representation of the YAML instead)
        obj (anything): the object to save
        folder (str): folder if not part of the filename
        die (bool): whether or not to raise an exception if saving an empty object
        indent (int): indentation to use for saved YAML
        keepnone (bool): allow :func:`sc.saveyaml(None) <saveyaml>` to return 'null' rather than raising an exception
        dumpall (bool): if True, treat a list input as separate YAML pages
        sanitizepath (bool): whether to sanitize the path prior to saving
        kwargs (dict): passed to ``yaml.dump()``

    Returns:
        The filename saved to

    **Examples**::

        yaml = {'foo':'bar', 'data':[1,2,3]}
        sc.saveyaml('my-file.yaml', yaml) # Save to file

        string = sc.saveyaml(obj=yaml) # Export to string
    '''
    import yaml # Optional import

    if dumpall: dump_func = yaml.dump_all
    else:       dump_func = yaml.dump

    if obj is None and not keepnone: # pragma: no cover
        errormsg = 'No object was supplied to saveyaml(), or the object was empty'
        if die: raise ValueError(errormsg)
        else:   print(errormsg)

    # Standard usage: dump to file
    if filename is not None:
        filename = makefilepath(filename=filename, folder=folder, sanitize=sanitizepath, makedirs=True)
        output = filename
        with open(filename, 'w') as f:
            dump_func(obj, f, **kwargs)

    # Alternate usage:
    else: # pragma: no cover
        output = dump_func(obj, **kwargs)

    return output


def jsonpickle(obj, tostring=False):
    '''
    Save any Python object to a JSON file using jsonpickle.

    Args:
        obj (any): the object to pickle as a JSON
        tostring (bool): whether to return a string (rather than the JSONified Python object)

    Returns:
        Either a string or a Python object for the JSON

    Wrapper for the jsonpickle library: https://jsonpickle.github.io/
    '''
    import jsonpickle as jp # Optional import
    import jsonpickle.ext.numpy as jsonpickle_numpy
    import jsonpickle.ext.pandas as jsonpickle_pandas
    jsonpickle_numpy.register_handlers()
    jsonpickle_pandas.register_handlers()

    if tostring:
        output = jp.dumps(obj)
    else:
        pickler = jp.pickler.Pickler()
        output = pickler.flatten(obj)

    return output


def jsonunpickle(json):
    ''' Use jsonpickle to restore an object (see jsonpickle()) '''
    import jsonpickle as jp
    import jsonpickle.ext.numpy as jsonpickle_numpy
    import jsonpickle.ext.pandas as jsonpickle_pandas
    jsonpickle_numpy.register_handlers()
    jsonpickle_pandas.register_handlers()

    if isinstance(json, str):
        output = jp.loads(json)
    else:
        unpickler = jp.unpickler.Unpickler()
        output = unpickler.restore(json)

    return output


##############################################################################
#%% Spreadsheet functions
##############################################################################

__all__ += ['Blobject', 'Spreadsheet', 'loadspreadsheet', 'savespreadsheet']


class Blobject(object):
    '''
    A wrapper for a binary file -- rarely used directly.

    So named because it's an object representing a blob.

    "source" is a specification of where to get the data from. It can be anything
    supported by Blobject.load() which are (a) a filename, which will get loaded,
    or (b) a io.BytesIO which will get dumped into this instance

    Alternatively, can specify ``blob`` which is a binary string that gets stored directly
    in the ``blob`` attribute
    '''

    def __init__(self, source=None, name=None, filename=None, blob=None):
        # Handle inputs
        if source   is None and filename is not None: source   = filename # Reset the source to be the filename, e.g. Spreadsheet(filename='foo.xlsx')
        if filename is None and scu.isstring(source):  filename = source   # Reset the filename to be the source, e.g. Spreadsheet('foo.xlsx')
        if name     is None and filename is not None: name     = os.path.basename(filename) # If not supplied, use the filename
        if blob is not None and source is not None: raise ValueError('Can initialize from either source or blob, but not both')

        # Define quantities
        self.name      = name # Name of the object
        self.filename  = filename # Filename (used as default for load/save)
        self.created   = scd.now() # When the object was created
        self.modified  = scd.now() # When it was last modified
        self.blob  = blob # The binary data
        self.bytes = None # The filestream representation of the binary data
        if source is not None: self.load(source)
        return


    def __repr__(self):
        return scp.prepr(self, skip=['blob','bytes'])


    def load(self, source=None):
        '''
        This function loads the spreadsheet from a file or object. If no input argument is supplied,
        then it will read self.bytes, assuming it exists.
        '''
        def read_bin(source):
            ''' Helper to read a binary stream '''
            source.flush()
            source.seek(0)
            output = source.read()
            return output

        def read_file(filename):
            ''' Helper to read an actual file '''
            filepath = makefilepath(filename=filename)
            self.filename = filename
            with open(filepath, mode='rb') as f:
                output = f.read()
            return output

        if isinstance(source, Path):  # If it's a path object, convert to string
            source = str(source)

        if source is None:
            if self.bytes is not None:
                self.blob = read_bin(self.bytes)
                self.bytes = None # Once read in, delete
            else: # pragma: no cover
                if self.filename is not None:
                    self.blob = read_file(self.filename)
                else:
                    print('Nothing to load: no source or filename supplied and self.bytes is empty.')
        else:
            if isinstance(source,io.BytesIO):
                self.blob = read_bin(source)
            elif scu.isstring(source):
                self.blob = read_file(source)
            else: # pragma: no cover
                errormsg = f'Input source must be type string (for a filename) or BytesIO, not {type(source)}'
                raise TypeError(errormsg)

        self.modified = scd.now()
        return


    def save(self, filename=None):
        ''' This function writes the spreadsheet to a file on disk. '''
        filepath = makefilepath(filename=filename, makedirs=True)
        with open(filepath, mode='wb') as f:
            f.write(self.blob)
        self.filename = filename
        print(f'Object saved to {filepath}.')
        return filepath


    def tofile(self, output=True):
        '''
        Return a file-like object with the contents of the file.

        This can then be used to open the workbook from memory without writing anything to disk e.g.

            - book = openpyxl.load_workbook(self.tofile())
            - book = xlrd.open_workbook(file_contents=self.tofile().read())
        '''
        bytesblob = io.BytesIO(self.blob)
        if output:
            return bytesblob
        else:
            self.bytes = bytesblob
            return


    def freshbytes(self):
        ''' Refresh the bytes object to accept new data '''
        self.bytes = io.BytesIO()
        return self.bytes



class Spreadsheet(Blobject):
    '''
    A class for reading and writing Excel files in binary format. No disk IO needs
    to happen to manipulate the spreadsheets with openpyxl (or xlrd or pandas).

    *New version 1.3.0:* Changed default from xlrd to openpyxl and added self.wb
    attribute to avoid the need to reload workbooks.

    **Examples**::


    '''

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.wb = None
        return


    def __getstate__(self): # pragma: no cover
        d = self.__dict__.copy() # Shallow copy
        d['wb'] = None
        return d


    def _reload_wb(self, reload=None):
        ''' Helper function to check if workbook is already loaded '''
        output = (not hasattr(self, 'wb')) or (self.wb is None) or reload
        return output


    def new(self, **kwargs):
        ''' Shortcut method to create a new openpyxl workbook '''
        import openpyxl # Optional import
        self.wb = openpyxl.Workbook(**kwargs)
        return self.wb


    def xlrd(self, reload=False, store=True, **kwargs): # pragma: no cover
        ''' Legacy method to load from xlrd '''
        if self._reload_wb(reload=reload):
            try:
                import xlrd # Optional import
            except ModuleNotFoundError as e:
                raise ModuleNotFoundError('The "xlrd" Python package is not available; please install manually') from e
            wb = xlrd.open_workbook(file_contents=self.tofile().read(), **kwargs)
        else:
            wb = self.wb

        if store:
            self.wb = wb
        return wb


    def openpyxl(self, reload=False, store=True, **kwargs): # pragma: no cover
        ''' Return a book as opened by openpyxl '''
        if self._reload_wb(reload=reload):
            import openpyxl # Optional import
            if self.blob is not None:
                self.tofile(output=False)
                wb = openpyxl.load_workbook(self.bytes, **kwargs) # This stream can be passed straight to openpyxl
            else:
                wb = openpyxl.Workbook(**kwargs)
        else:
            wb = self.wb

        if store:
            self.wb = wb
        return wb


    def openpyexcel(self, *args, **kwargs): # pragma: no cover
        ''' Legacy name for openpyxl() '''
        warnmsg = '''
Spreadsheet() no longer supports openpyexcel as of v1.3.1. To load using it anyway, you can manually do:

    %pip install openpyexcel
    import openpyexcel
    spreadsheet = sc.Spreadsheet()
    spreadsheet.wb = openpyexcel.Workbook(...)

Falling back to openpyxl, which is identical except for how cached cell values are handled.
'''
        warnings.warn(warnmsg, category=FutureWarning, stacklevel=2)
        return self.openpyxl(*args, **kwargs)


    def pandas(self, reload=False, store=True, **kwargs): # pragma: no cover
        ''' Return a book as opened by pandas '''

        if self._reload_wb(reload=reload):
            if self.blob is not None:
                self.tofile(output=False)
                wb = pd.ExcelFile(self.bytes, **kwargs)
            else:
                errormsg = 'For pandas, must load an existing workbook; use openpyxl to create a new workbook'
                raise FileNotFoundError(errormsg)
        else:
            wb = self.wb

        if store:
            self.wb = wb

        return wb


    def update(self, book): # pragma: no cover
        ''' Updated the stored spreadsheet with book instead '''
        self.tofile(output=False)
        book.save(self.freshbytes())
        self.load()
        return


    def _getsheet(self, sheetname=None, sheetnum=None):
        if   sheetname is not None: sheet = self.wb[sheetname]
        elif sheetnum  is not None: sheet = self.wb[self.wb.sheetnames[sheetnum]]
        else:                       sheet = self.wb.active
        return sheet


    def readcells(self, wbargs=None, *args, **kwargs):
        ''' Alias to loadspreadsheet() '''
        method = kwargs.pop('method', 'openpyxl')
        wbargs = scu.mergedicts(wbargs)
        f = self.tofile()
        kwargs['fileobj'] = f

        # Return the appropriate output
        cells = kwargs.pop('cells', None)

        # Read in sheetoutput (sciris dataframe object for xlrd, 2D numpy array for openpyxl).
        load_args = scu.mergedicts(dict(header=None), kwargs)
        if method == 'xlrd': # pragma: no cover
            sheetoutput = loadspreadsheet(*args, **load_args, method='xlrd')  # returns sciris dataframe object
        elif method == 'pandas':
            pandas_sheet = loadspreadsheet(*args, **load_args, method='pandas')
            sheetoutput = pandas_sheet.values
        elif method in ['openpyxl', 'openpyexcel']:
            wb_reader = self.openpyxl if method == 'openpyxl' else self.openpyexcel
            wb_reader(**wbargs)
            ws = self._getsheet(sheetname=kwargs.get('sheetname'), sheetnum=kwargs.get('sheetname'))
            rawdata = tuple(ws.rows)
            sheetoutput = np.empty(np.shape(rawdata), dtype=object)
            for r,rowdata in enumerate(rawdata):
                for c,val in enumerate(rowdata):
                    sheetoutput[r][c] = rawdata[r][c].value
        else: # pragma: no cover
            errormsg = f'Reading method not found; must be openpyxl or xlrd, not {method}'
            raise ValueError(errormsg)

        if cells is None:  # If no cells specified, return the whole sheet.
            return sheetoutput
        else:
            results = []
            for cell in cells:  # Loop over all cells
                rownum = cell[0]
                colnum = cell[1]
                if method in ['xlrd']:  # If we're using xlrd/pandas, reduce the row number by 1 # pragma: no cover
                    rownum -= 1
                results.append(sheetoutput[rownum][colnum])  # Grab and append the result at the cell.
            return results


    def writecells(self, cells=None, startrow=None, startcol=None, vals=None, sheetname=None, sheetnum=None, verbose=False, wbargs=None):
        '''
        Specify cells to write. Can supply either a list of cells of the same length
        as the values, or else specify a starting row and column and write the values
        from there.

        **Examples**::

            S = sc.Spreadsheet()
            S.writecells(cells=['A6','B7'], vals=['Cat','Dog']) # Method 1
            S.writecells(cells=[np.array([2,3])+i for i in range(2)], vals=['Foo', 'Bar']) # Method 2
            S.writecells(startrow=14, startcol=1, vals=np.random.rand(3,3)) # Method 3
            S.save('myfile.xlsx')
        '''
        # Load workbook
        if wbargs is None: wbargs = {}
        wb = self.openpyxl(**wbargs)
        if verbose: print(f'Workbook loaded: {wb}')

        # Get right worksheet
        ws = self._getsheet(sheetname=sheetname, sheetnum=sheetnum)
        if verbose: print(f'Worksheet loaded: {ws}')

        # Determine the cells
        if cells is not None: # A list of cells is supplied
            cells = scu.tolist(cells)
            vals  = scu.tolist(vals)
            if len(cells) != len(vals): # pragma: no cover
                errormsg = f'If using cells, cells and vals must have the same length ({len(cells)} vs. {len(vals)})'
                raise ValueError(errormsg)
            for cell,val in zip(cells,vals):
                try:
                    if scu.isstring(cell): # Handles e.g. cell='A1'
                        cellobj = ws[cell]
                    elif scu.checktype(cell, 'arraylike','number') and len(cell)==2: # Handles e.g. cell=(0,0)
                        cellobj = ws.cell(row=cell[0], column=cell[1])
                    else: # pragma: no cover
                        errormsg = f'Cell must be formatted as a label or row-column pair, e.g. "A1" or (3,5); not "{cell}"'
                        raise TypeError(errormsg)
                    if verbose: print(f'  Cell {cell} = {val}')
                    if isinstance(val, tuple): # pragma: no cover
                        cellobj.value = val[0]
                        cellobj.cached_value = val[1]
                    else:
                        cellobj.value = val
                except Exception as E: # pragma: no cover
                    errormsg = f'Could not write "{val}" to cell "{cell}": {repr(E)}'
                    raise RuntimeError(errormsg)
        else: # Cells aren't supplied, assume a matrix
            if startrow is None: startrow = 1 # Excel uses 1-based indexing
            if startcol is None: startcol = 1
            valarray = np.atleast_2d(np.array(vals, dtype=object))
            for i,rowvals in enumerate(valarray):
                row = startrow + i
                for j,val in enumerate(rowvals):
                    col = startcol + j
                    try:
                        key = f'row:{row} col:{col}'
                        ws.cell(row=row, column=col, value=val)
                        if verbose: print(f'  Cell {key} = {val}')
                    except Exception as E: # pragma: no cover
                        errormsg = f'Could not write "{val}" to {key}: {repr(E)}'
                        raise RuntimeError(errormsg)

        # Save
        wb.save(self.freshbytes())
        self.load()

        return


    def save(self, filename='spreadsheet.xlsx'):
        filepath = makefilepath(filename=filename, ext='xlsx', makedirs=True)
        Blobject.save(self, filepath)



def loadspreadsheet(filename=None, folder=None, fileobj=None, sheet=0, header=1, asdataframe=None, method='pandas', **kwargs):
    '''
    Load a spreadsheet as a dataframe or a list of lists.

    By default, an alias to :func:`pandas.read_excel()` with a header, but also supports loading
    via openpyxl or xlrd. Read from either a filename or a file object.

    Args:
        filename (str): filename or path to read
        folder (str): optional folder to use with the filename
        fileobj (obj): load from file object rather than path
        sheet (str/int/list): name or number of sheet(s) to use (default 0)
        asdataframe (bool): whether to return as a pandas/Sciris dataframe (default True)
        header (bool): whether the 0-th row is to be read as the header
        method (str): how to read (default 'pandas', other choices 'openpyxl' and 'xlrd')
        kwargs (dict): passed to pd.read_excel(), openpyxl(), etc.

    **Examples**::

        df = sc.loadspreadsheet('myfile.xlsx') # Alias to pd.read_excel(header=1)
        wb = sc.loadspreadsheet('myfile.xlsx', method='openpyxl') # Returns workbook
        data = sc.loadspreadsheet('myfile.xlsx', method='xlrd', asdataframe=False) # Returns raw data; requires xlrd

    *New version 1.3.0:* change default from xlrd to pandas; renamed sheetname and sheetnum arguments to sheet.
    '''

    # Handle path and sheet name/number
    fullpath = makefilepath(filename=filename, folder=folder)
    for key in ['sheetname', 'sheetnum', 'sheet_name']:
        sheet = kwargs.pop('sheetname', sheet)

    # Load using pandas
    if method == 'pandas':
        if fileobj is not None: fullpath = fileobj # Substitute here for reading
        data = pd.read_excel(fullpath, sheet_name=sheet, header=header, **kwargs)
        return data

    # Load using openpyxl
    elif method == 'openpyxl': # pragma: no cover
        spread = Spreadsheet(fullpath)
        wb = spread.openpyxl(**kwargs)
        return wb

    # Legacy method -- xlrd
    elif method == 'xlrd': # pragma: no cover
        try:
            import xlrd # Optional import

            # Handle inputs
            if asdataframe is None: asdataframe = True
            if isinstance(filename, io.BytesIO): fileobj = filename # It's actually a fileobj
            if fileobj is None:
                book = xlrd.open_workbook(fullpath)
            else:
                book = xlrd.open_workbook(file_contents=fileobj.read())

            if scu.isnumber(sheet):
                ws = book.sheet_by_index(sheet)
            else:
                ws = book.sheet_by_name(sheet)

            # Load the raw data
            rawdata = []
            for rownum in range(ws.nrows-header):
                rawdata.append(sco.odict())
                for colnum in range(ws.ncols):
                    if header: attr = ws.cell_value(0,colnum)
                    else:      attr = f'Column {colnum}'
                    attr = scu.uniquename(attr, namelist=rawdata[rownum].keys(), style='(%d)')
                    val = ws.cell_value(rownum+header,colnum)
                    try:
                        val = float(val) # Convert it to a number if possible
                    except:
                        try:    val = str(val)  # But give up easily and convert to a string (not Unicode)
                        except: pass # Still no dice? Fine, we tried
                    rawdata[rownum][str(attr)] = val

            # Convert to dataframe
            if asdataframe:
                cols = rawdata[0].keys()
                reformatted = []
                for oldrow in rawdata:
                    newrow = list(oldrow[:])
                    reformatted.append(newrow)
                dfdata = scdf.dataframe(cols=cols, data=reformatted)
                return dfdata

            # Or leave in the original format
            else:
                return rawdata
        except Exception as E:
            if isinstance(E, ModuleNotFoundError):
                errormsg = 'The "xlrd" Python package is not available; please install manually via "pip install xlrd==1.2.0"'
                raise ModuleNotFoundError(errormsg) from E
            elif isinstance(E, AttributeError):
                errormsg = '''
Warning! xlrd has been deprecated in Python 3.9 and can no longer read XLSX files.
If the error you got above is

    "AttributeError: 'ElementTree' object has no attribute 'getiterator'"

then this should fix it:

    import xlrd
    xlrd.xlsx.ensure_elementtree_imported(False, None)
    xlrd.xlsx.Element_has_iter = True

Then try again to load your Excel file.
'''
                raise AttributeError(errormsg) from E
            else:
                raise E

    else: # pragma: no cover
        errormsg = f'Method "{method}" not found: must be one of pandas, openpyxl, or xlrd'
        raise ValueError(errormsg)


def savespreadsheet(filename=None, data=None, folder=None, sheetnames=None, close=True,
                    workbook_args=None, formats=None, formatdata=None, verbose=False):
    '''
    Semi-simple function to save data nicely to Excel.

    Args:
        filename (str): Excel file to save to
        data (list/array): data to write to the spreadsheet
        folder (str): if supplied, merge with the filename to make a path
        sheetnames (list): if data is supplied as a list of arrays, save each entry to a different sheet
        close (bool): whether to close the workbook after saving
        workbook_args (dict): arguments passed to ``xlxwriter.Workbook()``
        formats (dict): a definition of different types of formatting (see examples below)
        formatdata (array): an array of which formats go where
        verbose (bool): whether to print progress

    **Examples**::

        import sciris as sc
        import pylab as pl

        # Simple example
        testdata1 = np.random.rand(8,3)
        sc.savespreadsheet(filename='test1.xlsx', data=testdata1)

        # Include column headers
        test2headers = [['A','B','C']] # Need double brackets to get right shape
        test2values = np.random.rand(8,3).tolist()
        testdata2 = test2headers + test2values
        sc.savespreadsheet(filename='test2.xlsx', data=testdata2)

        # Multiple sheets
        testdata3 = [np.random.rand(10,10), np.random.rand(20,5)]
        sheetnames = ['Ten by ten', 'Twenty by five']
        sc.savespreadsheet(filename='test3.xlsx', data=testdata3, sheetnames=sheetnames)

        # Supply data as an odict
        testdata4 = sc.odict([('First sheet', np.random.rand(6,2)), ('Second sheet', np.random.rand(3,3))])
        sc.savespreadsheet(filename='test4.xlsx', data=testdata4)

        # Include formatting
        nrows = 15
        ncols = 3
        formats = {
            'header':{'bold':True, 'bg_color':'#3c7d3e', 'color':'#ffffff'},
            'plain': {},
            'big':   {'bg_color':'#ffcccc'}
        }
        testdata5  = np.zeros((nrows+1, ncols), dtype=object) # Includes header row
        formatdata = np.zeros((nrows+1, ncols), dtype=object) # Format data needs to be the same size
        testdata5[0,:] = ['A', 'B', 'C'] # Create header
        testdata5[1:,:] = np.random.rand(nrows,ncols) # Create data
        formatdata[1:,:] = 'plain' # Format data
        formatdata[testdata5>0.7] = 'big' # Find "big" numbers and format them differently
        formatdata[0,:] = 'header' # Format header
        sc.savespreadsheet(filename='test5.xlsx', data=testdata5, formats=formats, formatdata=formatdata)

    *New version 2.0.0:* allow arguments to be passed to the ``Workbook``.
    '''
    workbook_args = scu.mergedicts({'nan_inf_to_errors': True}, workbook_args)
    try:
        import xlsxwriter # Optional import
    except ModuleNotFoundError as e: # pragma: no cover
        raise ModuleNotFoundError('The "xlsxwriter" Python package is not available; please install manually') from e
    fullpath = makefilepath(filename=filename, folder=folder, default='default.xlsx', makedirs=True)
    datadict   = sco.odict()
    formatdict = sco.odict()
    hasformats = (formats is not None) and (formatdata is not None)

    # Handle input arguments
    if isinstance(data, dict) and sheetnames is None: # pragma: no cover
        if verbose: print('Data is a dict, taking sheetnames from keys')
        sheetnames = data.keys()
        datadict   = sco.odict(data) # Use directly, converting to odict first
        if hasformats: formatdict = sco.odict(formatdata) #  NB, might be None but should be ok
    elif isinstance(data, dict) and sheetnames is not None: # pragma: no cover
        if verbose: print('Data is a dict, taking sheetnames from input')
        if len(sheetnames) != len(data):
            errormsg = f'If supplying data as a dict as well as sheetnames, length must match ({len(data)} vs {len(sheetnames)})'
            raise ValueError(errormsg)
        datadict   = sco.odict(data) # Use directly, but keep original sheet names
        if hasformats: formatdict = sco.odict(formatdata)
    elif not isinstance(data, dict):
        if sheetnames is None:
            if verbose: print('Data is a simple array')
            sheetnames = ['Sheet1']
            datadict[sheetnames[0]]   = data # Set it explicitly
            formatdict[sheetnames[0]] = formatdata # Set it explicitly -- NB, might be None but should be ok
        else: # pragma: no cover
            if verbose: print('Data is a list, taking matching sheetnames from inputs')
            if len(sheetnames) == len(data):
                for s,sheetname in enumerate(sheetnames):
                    datadict[sheetname] = data[s] # Assume there's a 1-to-1 mapping
                    if hasformats: formatdict[sheetname] = formatdata[s]
            else:
                errormsg = f'Unable to figure out how to match {len(sheetnames)} sheet names with data of length {len(data)}'
                raise ValueError(errormsg)
    else: # pragma: no cover
        errormsg = f'Cannot figure out how to handle data of type: {type(data)}'
        raise TypeError(errormsg) # This shouldn't happen!

    # Create workbook
    if verbose: print(f'Creating file {fullpath}')
    workbook = xlsxwriter.Workbook(fullpath, workbook_args)

    # Optionally add formats
    if formats is not None:
        if verbose: print(f'  Adding {len(formats)} formats')
        workbook_formats = dict()
        for formatkey,formatval in formats.items():
            workbook_formats[formatkey] = workbook.add_format(formatval)
    else:
        thisformat = workbook.add_format({}) # Plain formatting

    # Actually write the data
    for sheetname in datadict.keys():
        if verbose: print(f'  Creating sheet {sheetname}')
        sheetdata   = datadict[sheetname]
        if hasformats: sheetformat = formatdict[sheetname]
        worksheet = workbook.add_worksheet(sheetname)
        for r,row_data in enumerate(sheetdata):
            if verbose: print(f'    Writing row {r}/{len(sheetdata)}')
            for c,cell_data in enumerate(row_data):
                if verbose: print(f'      Writing cell {c}/{len(row_data)}')
                if hasformats:
                    thisformat = workbook_formats[sheetformat[r][c]] # Get the actual format
                try:
                    if not scu.isnumber(cell_data):
                        cell_data = str(cell_data) # Convert everything except numbers to strings -- use formatting to handle the rest
                    worksheet.write(r, c, cell_data, thisformat) # Write with or without formatting
                except Exception as E: # pragma: no cover
                    errormsg = f'Could not write cell [{r},{c}] data="{cell_data}", format="{thisformat}": {str(E)}'
                    raise RuntimeError(errormsg)

    # Either close the workbook and write to file, or return it for further working
    if close:
        if verbose: print(f'Saving file {filename} and closing')
        workbook.close()
        return fullpath
    else: # pragma: no cover
        if verbose: print('Returning workbook')
        return workbook



##############################################################################
#%% Pickling support methods
##############################################################################

__all__ += ['Failed', 'Empty']


class Failed(object):
    ''' An empty class to represent a failed object loading '''
    failure_info = sco.odict()

    def __init__(self, *args, **kwargs):
        pass

    def __repr__(self): # pragma: no cover
        output = scp.objrepr(self) # This does not include failure_info since it's a class attribute
        output += self.showfailures(verbose=False, tostring=True)
        return output

    def showfailures(self, verbose=True, tostring=False):
        output = ''
        for f,failure in self.failure_info.enumvals():
            output += f'\nFailure {f+1} of {len(self.failure_info)}:\n'
            output += f'Module: {failure["module"]}\n'
            output += f'Class: {failure["class"]}\n'
            output += f'Error: {failure["error"]}\n'
            if verbose: # pragma: no cover
                output += '\nTraceback:\n'
                output += f"\n{failure['exception']}"
                output += '\n\n'
        if tostring: # pragma: no cover
            return output
        else:
            print(output)
            return


class Empty(object): # pragma: no cover
    ''' Another empty class to represent a failed object loading, but do not proceed with setstate '''

    def __init__(self, *args, **kwargs):
        pass

    def __setstate__(self, state):
        pass


class UniversalFailed(Failed): # pragma: no cover
    ''' A universal failed object class, that preserves as much data as possible '''

    def __init__(self, *args, **kwargs):
        if args:
            self.args = args
        if kwargs:
            self.kwargs = kwargs
        self.__set_empty()
        return

    def __set_empty(self):
        if not hasattr(self, 'state'):
            self.state = {}
        if not hasattr(self, 'dict'):
            self.dict = {}
        return

    def __repr__(self):
        output = scp.objrepr(self) # This does not include failure_info since it's a class attribute
        return output

    def __setstate__(self, state):
        try:
            self.__dict__.update(state) # NB, does not support slots
        except:
            self.__set_empty()
            self.state = state
        return

    def __len__(self):
        return len(self.dict) if hasattr(self, 'dict') else 0

    def __getitem__(self, key):
        return self.dict[key]

    def __setitem__(self, key, value):
        self.dict[key] = value
        return

    def disp(self, *args, **kwargs):
        return scp.pr(self, *args, **kwargs)


def makefailed(module_name=None, name=None, error=None, exception=None, universal=False):
    ''' Create a class -- not an object! -- that contains the failure info for a pickle that failed to load '''
    base = UniversalFailed if universal else Failed
    key = f'Failure {len(base.failure_info)+1}'
    base.failure_info[key] = sco.odict()
    base.failure_info[key]['module']    = module_name
    base.failure_info[key]['class']     = name
    base.failure_info[key]['error']     = error
    base.failure_info[key]['exception'] = exception
    return base


def _remap_module(remapping, module_name, name):
    ''' Use a remapping dictionary to try to load a module from a different location ''' 
    key = f'{module_name}.{name}'
    obj = remapping.get(key) # If the user has supplied the module directly
    if isinstance(obj, str): # Split a string into a tuple, e.g. 'foo.bar.Cat' to ('foo.bar', 'Cat')
        obj = tuple(obj.rsplit('.', 1))
    if obj is None or (isinstance(obj, tuple) and len(obj)==2): # Either it's not in the remapping, or it's a tuple
        if obj is None: # Key not in remapping, just use regular
            load_module = module_name
            load_name = name
        else: # It's a tuple of names, process # pragma: no cover
            load_module = obj[0]
            load_name = obj[1]
        module = importlib.import_module(load_module)
        obj = getattr(module, load_name) # pragma: no cover
    return obj


class _LoadsInterface():
    # Add a .loads method to unpicklers, with support for Sciris remapping
    @classmethod
    def loads(cls, string, remapping, **kwargs):
        unpickler = cls(io.BytesIO(string), remapping=remapping, **kwargs)
        return unpickler.load()

class _RobustUnpickler(dill.Unpickler, _LoadsInterface):
    ''' Try to import an object, and if that fails, return a Failed object rather than crashing '''

    def __init__(self, bytesio, fix_imports=True, encoding="latin1", errors="ignore", remapping=None):
        super().__init__(bytesio, fix_imports=fix_imports, encoding=encoding, errors=errors)
        self.remapping = scu.mergedicts(remapping)
        return

    def find_class(self, module_name, name, verbose=True):
        obj = _remap_module(self.remapping, module_name, name)
        return obj


class _RenamingUnpickler(_RobustUnpickler, _LoadsInterface):
    ''' Like RobustUnpickler but with automatic population of 'remapping' with known fixes '''

    known_fixes = {
        ('pandas.core.indexes.numeric', 'Int64Index'): {'pandas.core.indexes.numeric.Int64Index':'pandas.core.indexes.api.Index'},
        ('pandas.core.indexes.numeric', 'Float64Index'): {'pandas.core.indexes.numeric.Float64Index':'pandas.core.indexes.api.Index'},
    }

    def find_class(self, module_name, name, last_error = None, verbose=True):

        try:
            return super().find_class(module_name, name, verbose)
        except ModuleNotFoundError as E:
            if last_error == E:
                raise E # We already tried to fix the error and it didn't work

            if (module_name, name) in self.known_fixes:
                self.remapping.update(self.known_fixes[(module_name, name)])
            elif module_name in self.known_fixes:
                self.remapping.update(self.known_fixes[module_name])
            else:
                raise E

            warnmsg = f'Fixing known unpickling deprecation "{str(E)}"'
            warnings.warn(warnmsg, category=UserWarning, stacklevel=2)
            return self.find_class(module_name, name, E, verbose)

class _UltraRobustUnpickler(dill.Unpickler, _LoadsInterface): # pragma: no cover
    ''' If all else fails, just make a default object '''

    def __init__(self, bytesio, *args, fix_imports=True, encoding="latin1", errors="ignore", 
                 remapping=None, unpicklingerrors=None):
        super().__init__(bytesio, fix_imports=fix_imports, encoding=encoding, errors=errors)
        self.remapping = scu.mergedicts(remapping)
        self.unpicklingerrors = unpicklingerrors if unpicklingerrors is not None else []
        return

    def find_class(self, module_name, name):
        ''' Ignore all attempts to use the actual class and always make a UniversalFailed class '''
        try:
            obj = _remap_module(self.remapping, module_name, name)
        except Exception as E:
            obj = makefailed(module_name=module_name, name=name, error=str(E), exception=E, universal=True)
        return obj


def _unpickler(string=None, die=False, verbose=False, remapping=None, method='pickle', auto_remap=True, **kwargs):
    ''' Not invoked directly; used as a helper function for saveobj/loadobj '''

    # Sanitize kwargs, since wrapped in try-except statements otherwise
    if kwargs: # pragma: no cover
        valid_kwargs = ['fix_imports', 'encoding', 'errors', 'buffers', 'ignore']
        for k in kwargs:
            if k not in valid_kwargs:
                errormsg = f'Keyword "{k}" is not a valid keyword: {scu.strjoin(valid_kwargs)}'
                raise ValueError(errormsg)

    methods = {
        'Pickle': pkl.loads,
        'Pandas': lambda string, **kwargs: pd.read_pickle(io.BytesIO(string), **kwargs),
        'Dill': dill.loads,
        'Encoded Pickle': lambda string, **kwargs: pkl.loads(string, encoding='latin1', **kwargs),
        'Robust': lambda string, **kwargs: _RobustUnpickler.loads(string, remapping, **kwargs),
        'Robust (renaming)': lambda string, **kwargs: _RenamingUnpickler.loads(string, remapping, **kwargs),
        'Ultrarobust': lambda string, **kwargs: _UltraRobustUnpickler.loads(string, remapping, **kwargs),
    }

    # Methods that allow for loading an object without any failed instances (if die=True)
    if method == 'pickle':
        unpicklers = ['Pickle','Pandas','Encoded Pickle']
    elif method == 'dill':
        unpicklers = ['Dill']
    else:
        unpicklers = ['Pickle','Pandas','Dill','Encoded Pickle']

    # If permitted, return an object that encountered errors in the loading process and therefore may not be valid
    # Such an object might require further fixes to be made by the user
    if not die:
        unpicklers += ['Robust (renaming)' if auto_remap else 'Robust', 'Ultrarobust']

    errors = {}
    obj = None

    for unpickler in unpicklers:
        try:
            obj = methods[unpickler](string, **kwargs)
            break
        except Exception as E:
            errors[unpickler] = str(E)
            if verbose: print(f'{unpickler} failed ({str(E)})')

    if obj is None:
        errormsg = 'All available unpickling methods failed: ' + '\n'.join([f'{k}: {v}' for k,v in errors.items()])
        raise Exception(errormsg)
    elif isinstance(obj, Failed):
        print('Warning, the following errors were encountered during unpickling:')
        obj.showfailures(verbose=False)

    return obj


def _savepickle(fileobj=None, obj=None, protocol=None, **kwargs):
        ''' Use pickle to do the salty work. '''
        if protocol is None:
            protocol = 4 # Use protocol 4 for backwards compatibility
        fileobj.write(pkl.dumps(obj, protocol=protocol, **kwargs))
        return


def _savedill(fileobj=None, obj=None, **kwargs): # pragma: no cover
    ''' Use dill to do the sour work (note: this function is not actively maintained) '''
    fileobj.write(dill.dumps(obj, protocol=-1, **kwargs))
    return
